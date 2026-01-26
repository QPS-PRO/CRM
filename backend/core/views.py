from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook
from datetime import datetime
import re
from .models import Parent, Student, Branch, Grade, UserProfile, UserRole
from .serializers import ParentSerializer, StudentSerializer, BranchSerializer, UserSerializer, UserCreateSerializer


class BranchViewSet(viewsets.ModelViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'address']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']


class ParentViewSet(viewsets.ModelViewSet):
    queryset = Parent.objects.prefetch_related('students').all()
    serializer_class = ParentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email', 'phone_number', 'students__student_id']
    filterset_fields = ['email']
    ordering_fields = ['created_at', 'first_name', 'last_name']
    ordering = ['-created_at']

    @action(detail=False, methods=['get'])
    def all(self, request):
        """Get all parents without pagination"""
        parents = self.get_queryset()
        serializer = self.get_serializer(parents, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """Bulk upload parents from Excel file"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            workbook = load_workbook(file, data_only=True)
            sheet = workbook.active
            
            # Get header row
            headers = [cell.value.lower().strip() if cell.value else '' for cell in sheet[1]]
            
            # Expected columns: first_name, last_name, email, phone_number, address, student_id
            required_columns = ['first_name', 'last_name', 'email', 'student_id']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                return Response({
                    'error': f'Missing required columns: {", ".join(missing_columns)}',
                    'found_columns': headers
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create column index mapping
            col_map = {header: idx for idx, header in enumerate(headers)}
            
            created_count = 0
            updated_count = 0
            linked_count = 0
            errors = []
            
            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                    try:
                        # Skip empty rows
                        if not any(cell.value for cell in row):
                            continue
                        
                        # Extract data
                        first_name = str(row[col_map['first_name']].value or '').strip()
                        last_name = str(row[col_map['last_name']].value or '').strip()
                        email = str(row[col_map['email']].value or '').strip()
                        student_ids_str = str(row[col_map['student_id']].value or '').strip()
                        
                        # Get optional fields
                        phone_number = ''
                        if 'phone_number' in col_map:
                            phone_number = str(row[col_map['phone_number']].value or '').strip()
                        
                        address = ''
                        if 'address' in col_map:
                            address = str(row[col_map['address']].value or '').strip()
                        
                        # Validation
                        if not first_name or not last_name:
                            errors.append(f'Row {row_idx}: Missing required fields (first_name or last_name)')
                            continue
                        
                        if not email:
                            errors.append(f'Row {row_idx}: Missing required field (email)')
                            continue
                        
                        # Validate email format
                        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                        if not re.match(email_pattern, email):
                            errors.append(f'Row {row_idx}: Invalid email format: {email}')
                            continue
                        
                        # Parse student IDs (can be comma or semicolon separated)
                        student_ids = []
                        if student_ids_str:
                            # Split by comma or semicolon
                            ids = re.split(r'[,;]', student_ids_str)
                            student_ids = [id.strip() for id in ids if id.strip()]
                        
                        # Create or update parent
                        parent, created = Parent.objects.update_or_create(
                            email=email,
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'phone_number': phone_number,
                                'address': address,
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                        
                        # Link to students
                        if student_ids:
                            students = Student.objects.filter(student_id__in=student_ids)
                            if students.exists():
                                parent.students.add(*students)
                                linked_count += students.count()
                            else:
                                errors.append(f'Row {row_idx}: No students found with IDs: {", ".join(student_ids)}')
                            
                    except Exception as e:
                        errors.append(f'Row {row_idx}: {str(e)}')
                        continue
            
            return Response({
                'message': 'Bulk upload completed',
                'created': created_count,
                'updated': updated_count,
                'students_linked': linked_count,
                'errors': errors[:50],  # Limit errors to first 50
                'total_errors': len(errors)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Error processing file: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.prefetch_related('parents').all()
    serializer_class = StudentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'student_id', 'parents__first_name', 'parents__last_name']
    filterset_fields = ['grade', 'level', 'class_name', 'gender', 'is_active', 'parents', 'branch']
    ordering_fields = ['created_at', 'first_name', 'last_name', 'student_id']
    ordering = ['-created_at']

    @action(detail=False, methods=['get'])
    def by_grade(self, request):
        """Get students grouped by grade"""
        grade = request.query_params.get('grade')
        if grade:
            students = self.queryset.filter(grade=grade, is_active=True)
            serializer = self.get_serializer(students, many=True)
            return Response(serializer.data)
        return Response({'error': 'grade parameter is required'}, status=400)

    @action(detail=False, methods=['get'])
    def classes(self, request):
        """Get unique class names filtered by branch, grade, and optionally level"""
        queryset = self.queryset.filter(is_active=True)
        
        branch_id = request.query_params.get('branch')
        grade = request.query_params.get('grade')
        level = request.query_params.get('level')
        
        if branch_id:
            try:
                branch = Branch.objects.get(id=branch_id)
                queryset = queryset.filter(branch=branch)
            except Branch.DoesNotExist:
                pass
        
        if grade:
            queryset = queryset.filter(grade=grade)
        
        if level:
            queryset = queryset.filter(level=level)
        
        # Get distinct class names, excluding empty/null values
        class_names = queryset.values_list('class_name', flat=True)
        # Filter out empty/null values and strip whitespace
        class_names = [name.strip() for name in class_names if name and name.strip()]
        # Use set to remove duplicates, then convert back to sorted list
        class_names = sorted(set(class_names))
        
        return Response(class_names)

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """Bulk upload students from Excel file"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            workbook = load_workbook(file, data_only=True)
            sheet = workbook.active
            
            # Get header row
            headers = [cell.value.lower().strip() if cell.value else '' for cell in sheet[1]]
            
            # Expected columns: first_name, last_name, student_id, grade, level, class, gender, date_of_birth, branch
            required_columns = ['first_name', 'last_name', 'student_id', 'grade', 'gender', 'date_of_birth', 'branch']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                return Response({
                    'error': f'Missing required columns: {", ".join(missing_columns)}',
                    'found_columns': headers
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create column index mapping
            col_map = {header: idx for idx, header in enumerate(headers)}
            
            created_count = 0
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                    try:
                        # Skip empty rows
                        if not any(cell.value for cell in row):
                            continue
                        
                        # Extract data
                        first_name = str(row[col_map['first_name']].value or '').strip()
                        last_name = str(row[col_map['last_name']].value or '').strip()
                        student_id = str(row[col_map['student_id']].value or '').strip()
                        grade = str(row[col_map['grade']].value or '').strip().upper()
                        gender = str(row[col_map['gender']].value or '').strip().upper()
                        date_of_birth_str = row[col_map['date_of_birth']].value
                        branch_name = str(row[col_map['branch']].value or '').strip()
                        
                        # Get optional fields
                        level = None
                        if 'level' in col_map:
                            level_val = row[col_map['level']].value
                            if level_val:
                                try:
                                    level = int(float(str(level_val)))
                                    if level < 1 or level > 12:
                                        level = None
                                except (ValueError, TypeError):
                                    level = None
                        
                        class_name = ''
                        if 'class' in col_map:
                            class_name = str(row[col_map['class']].value or '').strip()
                        
                        # Validation
                        if not first_name or not last_name or not student_id:
                            errors.append(f'Row {row_idx}: Missing required fields (first_name, last_name, or student_id)')
                            continue
                        
                        valid_grades = [g[0] for g in Grade.choices]
                        if grade not in valid_grades:
                            errors.append(f'Row {row_idx}: Invalid grade "{grade}". Must be one of: {", ".join(valid_grades)}')
                            continue
                        
                        if gender not in ['M', 'F']:
                            errors.append(f'Row {row_idx}: Invalid gender "{gender}". Must be M or F')
                            continue
                        
                        # Parse date of birth
                        date_of_birth = None
                        if isinstance(date_of_birth_str, datetime):
                            date_of_birth = date_of_birth_str.date()
                        elif isinstance(date_of_birth_str, str):
                            # Try different date formats
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                try:
                                    date_of_birth = datetime.strptime(date_of_birth_str.strip(), fmt).date()
                                    break
                                except ValueError:
                                    continue
                        elif date_of_birth_str:
                            # Try to convert Excel date number
                            try:
                                date_of_birth = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_of_birth_str) - 2).date()
                            except (ValueError, TypeError):
                                pass
                        
                        if not date_of_birth:
                            errors.append(f'Row {row_idx}: Invalid date_of_birth format')
                            continue
                        
                        # Get or create branch
                        branch, _ = Branch.objects.get_or_create(
                            name=branch_name,
                            defaults={'address': ''}
                        )
                        
                        # Create or update student
                        student, created = Student.objects.update_or_create(
                            student_id=student_id,
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'grade': grade,
                                'level': level,
                                'class_name': class_name,
                                'gender': gender,
                                'date_of_birth': date_of_birth,
                                'branch': branch,
                                'is_active': True,
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                            
                    except Exception as e:
                        errors.append(f'Row {row_idx}: {str(e)}')
                        continue
            
            return Response({
                'message': 'Bulk upload completed',
                'created': created_count,
                'updated': updated_count,
                'errors': errors[:50],  # Limit errors to first 50
                'total_errors': len(errors)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Error processing file: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """Login endpoint"""
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response(
            {'error': 'Username and password are required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(request, username=username, password=password)
    
    if user is not None:
        if user.is_active:
            login(request, user)
            # Get or create profile to get role
            profile, _ = UserProfile.objects.get_or_create(
                user=user,
                defaults={'role': UserRole.ADMIN if (user.is_superuser or user.is_staff) else UserRole.VIEWER}
            )
            return Response({
                'message': 'Login successful',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'is_staff': user.is_staff,
                    'is_superuser': user.is_superuser,
                    'role': profile.role,
                }
            })
        else:
            return Response(
                {'error': 'User account is disabled'},
                status=status.HTTP_403_FORBIDDEN
            )
    else:
        return Response(
            {'error': 'Invalid username or password'},
            status=status.HTTP_401_UNAUTHORIZED
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    """Logout endpoint"""
    logout(request)
    return Response({'message': 'Logout successful'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user_view(request):
    """Get current authenticated user"""
    user = request.user
    # Get or create profile
    profile, _ = UserProfile.objects.get_or_create(
        user=user,
        defaults={'role': UserRole.ADMIN if (user.is_superuser or user.is_staff) else UserRole.VIEWER}
    )
    return Response({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser,
        'role': profile.role,
    })


class UserViewSet(viewsets.ModelViewSet):
    """ViewSet for user management (admin only)"""
    queryset = User.objects.all().select_related('profile')
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['username', 'email']
    ordering_fields = ['username', 'date_joined']
    ordering = ['-date_joined']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer
    
    def get_permissions(self):
        """Only admins can manage users"""
        if self.action in ['list', 'create', 'retrieve', 'update', 'partial_update', 'destroy']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [AllowAny]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        """Filter queryset based on permissions"""
        user = self.request.user
        # Only admins can see all users
        if user.is_authenticated:
            profile, _ = UserProfile.objects.get_or_create(
                user=user,
                defaults={'role': UserRole.ADMIN if (user.is_superuser or user.is_staff) else UserRole.VIEWER}
            )
            if profile.is_admin:
                return self.queryset
        return User.objects.none()
    
    def create(self, request, *args, **kwargs):
        """Create a new user (admin only)"""
        user = request.user
        if not user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        profile, _ = UserProfile.objects.get_or_create(
            user=user,
            defaults={'role': UserRole.ADMIN if (user.is_superuser or user.is_staff) else UserRole.VIEWER}
        )
        
        if not profile.is_admin:
            return Response({'error': 'Only admins can create users'}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user_instance = serializer.save()
        
        # Return the created user with role
        response_serializer = UserSerializer(user_instance)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)
    
    def destroy(self, request, *args, **kwargs):
        """Delete a user (admin only, cannot delete self)"""
        user = request.user
        if not user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        profile, _ = UserProfile.objects.get_or_create(
            user=user,
            defaults={'role': UserRole.ADMIN if (user.is_superuser or user.is_staff) else UserRole.VIEWER}
        )
        
        if not profile.is_admin:
            return Response({'error': 'Only admins can delete users'}, status=status.HTTP_403_FORBIDDEN)
        
        instance = self.get_object()
        if instance.id == user.id:
            return Response({'error': 'Cannot delete your own account'}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().destroy(request, *args, **kwargs)

